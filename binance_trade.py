"""
Binance Futures 下單模組 (binance_trade.py)
- 支援 Testnet / 正式環境切換
- 市價下單 + 分開掛 SL/TP
- 部分平倉（TP1 10%）
- Excel 交易紀錄
"""
import os
import time
import math
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from binance.um_futures import UMFutures
from telegram_notify import send_telegram_message

load_dotenv()

# ── API 設定 ──────────────────────────────────────────────────
api_key = os.getenv("BINANCE_API_KEY", "")
api_secret = os.getenv("BINANCE_API_SECRET", "")
is_testnet = os.getenv("BINANCE_TESTNET", "true").lower() == "true"

TESTNET_URL = "https://testnet.binancefuture.com"
PRODUCTION_URL = "https://fapi.binance.com"

# ── 倉位與風控參數 ───────────────────────────────────────────
SYMBOL = os.getenv("SYMBOL", "BTCUSDT")
MARGIN_PER_TRADE = float(os.getenv("MARGIN_PER_TRADE", 100))
LEVERAGE = int(os.getenv("LEVERAGE", 20))
COOLDOWN_SECONDS = int(os.getenv("COOLDOWN_SECONDS", 60))

# ── 時間校正（修正本機與伺服器時間差）────────────────────────
import binance.lib.utils as _binance_utils
_original_get_timestamp = _binance_utils.get_timestamp
_time_offset = 0  # ms


def sync_time():
    """同步本機與 Binance 伺服器的時間差，patch SDK 的 get_timestamp"""
    global _time_offset
    import requests as _req
    base_url = TESTNET_URL if is_testnet else PRODUCTION_URL
    try:
        resp = _req.get(f"{base_url}/fapi/v1/time", timeout=5)
        server_time = resp.json()["serverTime"]
        local_time = int(time.time() * 1000)
        _time_offset = server_time - local_time
        if abs(_time_offset) > 500:
            print(f"Time offset: {_time_offset}ms (synced)")
    except Exception as e:
        print(f"Time sync failed: {e}")


def _patched_get_timestamp():
    return int(time.time() * 1000) + _time_offset

_binance_utils.get_timestamp = _patched_get_timestamp

# 也 patch binance.api 模組裡的本地引用
import binance.api as _binance_api
_binance_api.get_timestamp = _patched_get_timestamp


# ── 建立 Binance Session ─────────────────────────────────────
def new_session():
    base_url = TESTNET_URL if is_testnet else PRODUCTION_URL
    sync_time()
    return UMFutures(key=api_key, secret=api_secret, base_url=base_url)

client = new_session()
last_trade_time = {}
_session_created_at = time.time()
_SESSION_MAX_AGE = 1800  # 每 30 分鐘重建 session（防止連線過期）


def _ensure_session():
    """確保 session 有效，過期則重建"""
    global client, _session_created_at
    if time.time() - _session_created_at > _SESSION_MAX_AGE:
        client = new_session()
        _session_created_at = time.time()

# ── 快取商品規格 ─────────────────────────────────────────────
_symbol_info_cache = {}


def get_symbol_info(symbol=None):
    """查詢商品規格（tickSize, qtyStep, minQty），有快取"""
    symbol = symbol or SYMBOL
    if symbol in _symbol_info_cache:
        return _symbol_info_cache[symbol]

    global client
    try:
        info = client.exchange_info()
        for s in info["symbols"]:
            if s["symbol"] == symbol:
                tick_size = None
                qty_step = None
                min_qty = None
                for f in s["filters"]:
                    if f["filterType"] == "PRICE_FILTER":
                        tick_size = float(f["tickSize"])
                    elif f["filterType"] == "LOT_SIZE":
                        qty_step = float(f["stepSize"])
                        min_qty = float(f["minQty"])

                if tick_size and qty_step and min_qty:
                    _symbol_info_cache[symbol] = (tick_size, qty_step, min_qty)
                    return tick_size, qty_step, min_qty

        print(f"Symbol {symbol} not found in exchange info")
        return 0.10, 0.001, 0.001
    except Exception as e:
        print(f"get_symbol_info error: {e}")
        client = new_session()
        return 0.10, 0.001, 0.001


def get_available_balance():
    """查詢 USDT 可用餘額"""
    global client
    _ensure_session()
    try:
        account = client.account()
        for asset in account["assets"]:
            if asset["asset"] == "USDT":
                return float(asset["availableBalance"])
        return 0.0
    except Exception as e:
        print(f"get_available_balance error: {e}")
        send_telegram_message(f"查詢餘額失敗: {e}")
        client = new_session()
        return 0.0


def get_positions(symbol=None):
    """查詢當前持倉"""
    symbol = symbol or SYMBOL
    global client
    _ensure_session()
    try:
        positions = client.get_position_risk(symbol=symbol)
        result = []
        for p in positions:
            amt = float(p["positionAmt"])
            if amt != 0:
                result.append({
                    "symbol": p["symbol"],
                    "side": "long" if amt > 0 else "short",
                    "size": abs(amt),
                    "entry_price": float(p["entryPrice"]),
                    "unrealized_pnl": float(p["unRealizedProfit"]),
                    "leverage": int(p.get("leverage", LEVERAGE)),
                    "mark_price": float(p["markPrice"]),
                })
        return result
    except Exception as e:
        print(f"get_positions error: {e}")
        client = new_session()
        return []


def set_leverage(symbol=None, leverage=None):
    """設定槓桿"""
    symbol = symbol or SYMBOL
    leverage = leverage or LEVERAGE
    global client
    try:
        client.change_leverage(symbol=symbol, leverage=leverage)
        print(f"Leverage set to {leverage}x for {symbol}")
    except Exception as e:
        # 已經是該槓桿就會報錯，忽略
        if "No need to change" not in str(e):
            print(f"set_leverage: {e}")


# ── 精度工具 ─────────────────────────────────────────────────
def round_to_tick(price, tick_size):
    decimals = max(0, -int(math.floor(math.log10(tick_size))))
    return round(round(price / tick_size) * tick_size, decimals)


def round_to_lot(qty, qty_step, min_qty):
    decimals = max(0, -int(math.floor(math.log10(qty_step))))
    qty = round(round(qty / qty_step) * qty_step, decimals)
    return max(qty, min_qty)


# ── 下單 ─────────────────────────────────────────────────────
def place_order(symbol, side, qty=None, stop_loss=None, take_profit=None,
                reduce_only=False, strategy_id="v3"):
    """
    市價下單。
    side: "BUY" or "SELL"
    qty: 下單數量（BTC），None 則自動計算
    stop_loss: 止損價格
    take_profit: 止盈價格
    reduce_only: True = 平倉單
    """
    global client, last_trade_time

    now = time.time()
    tick_size, qty_step, min_qty = get_symbol_info(symbol)

    # 冷卻檢查（僅開倉單）
    if not reduce_only:
        cooldown_key = (strategy_id, symbol)
        if cooldown_key in last_trade_time and now - last_trade_time[cooldown_key] < COOLDOWN_SECONDS:
            print(f"Cooldown active for {strategy_id}/{symbol}")
            return None

    # 自動計算數量（含重試）
    if qty is None:
        mark_price = 0
        for _attempt in range(3):
            try:
                mark_price = float(client.mark_price(symbol=symbol)["markPrice"])
                if mark_price > 0:
                    break
            except Exception:
                time.sleep(1)
        if mark_price <= 0:
            print("Cannot get mark price after 3 attempts")
            return None
        notional = MARGIN_PER_TRADE * LEVERAGE
        qty = notional / mark_price

    qty = round_to_lot(qty, qty_step, min_qty)

    try:
        params = {
            "symbol": symbol,
            "side": side.upper(),
            "type": "MARKET",
            "quantity": qty,
        }
        if reduce_only:
            params["reduceOnly"] = True

        res = client.new_order(**params)
        order_id = res.get("orderId", "?")
        avg_price = float(res.get("avgPrice", 0))

        if reduce_only:
            msg = (f"<b>[平倉] {side} {symbol}</b>\n"
                   f"數量: {qty}  價格: {avg_price:.2f}\n"
                   f"OrderID: {order_id}")
        else:
            notional = qty * avg_price if avg_price > 0 else MARGIN_PER_TRADE * LEVERAGE
            margin_used = notional / LEVERAGE if LEVERAGE > 0 else notional
            sl_text = f"\nSL: {stop_loss:.2f}" if stop_loss else ""
            tp_text = f"  TP: {take_profit:.2f}" if take_profit else ""
            msg = (f"<b>[開倉] {side} {symbol}</b>\n"
                   f"數量: {qty}  價格: {avg_price:.2f}\n"
                   f"名義: {notional:.2f} USDT  保證金: {margin_used:.2f} USDT"
                   f"{sl_text}{tp_text}\n"
                   f"OrderID: {order_id}")
        print(msg)
        send_telegram_message(msg)

        if not reduce_only:
            last_trade_time[(strategy_id, symbol)] = now

        # 掛 SL/TP（僅開倉單）
        if not reduce_only:
            if stop_loss:
                _place_stop_order(symbol, side, qty, stop_loss, "STOP_MARKET", tick_size)
            if take_profit:
                _place_stop_order(symbol, side, qty, take_profit, "TAKE_PROFIT_MARKET", tick_size)

        return res

    except Exception as e:
        print(f"place_order error: {e}")
        send_telegram_message(f"<b>[下單失敗]</b> {side} {symbol}\n{e}")
        client = new_session()
        return None


def _place_stop_order(symbol, entry_side, qty, price, order_type, tick_size):
    """掛止損/止盈單（使用 Algo Order API）"""
    global client
    close_side = "SELL" if entry_side.upper() == "BUY" else "BUY"
    price = round_to_tick(price, tick_size)

    try:
        params = {
            "algoType": "CONDITIONAL",
            "symbol": symbol,
            "side": close_side,
            "type": order_type,
            "triggerPrice": str(price),
            "closePosition": "true",
        }
        res = client.sign_request("POST", "/fapi/v1/algoOrder", params)
        label = "SL" if "STOP" in order_type else "TP"
        algo_id = res.get("algoId", "?")
        print(f"  {label} algo order placed: {close_side} at {price} (algoId={algo_id})")
        return res
    except Exception as e:
        print(f"  Algo stop order error ({order_type}): {e}")
        return None


def close_position(symbol=None, side=None):
    """平倉指定方向的持倉"""
    symbol = symbol or SYMBOL
    positions = get_positions(symbol)

    for pos in positions:
        if side and pos["side"] != side:
            continue
        close_side = "SELL" if pos["side"] == "long" else "BUY"
        place_order(symbol, close_side, qty=pos["size"], reduce_only=True)


def cancel_all_orders(symbol=None):
    """取消所有掛單（含 algo orders）"""
    symbol = symbol or SYMBOL
    global client

    # 取消一般掛單
    try:
        client.cancel_open_orders(symbol=symbol)
        print(f"All open orders cancelled for {symbol}")
    except Exception as e:
        if "No open orders" not in str(e):
            print(f"cancel_all_orders: {e}")

    # 取消 algo orders
    try:
        algo_orders = client.sign_request("GET", "/fapi/v1/openAlgoOrders", {
            "symbol": symbol,
            "algoType": "CONDITIONAL",
        })
        for order in algo_orders:
            algo_id = order["algoId"]
            client.sign_request("DELETE", "/fapi/v1/algoOrder", {"algoId": algo_id})
            print(f"  Cancelled algo order {algo_id}")
    except Exception as e:
        print(f"cancel algo orders error: {e}")


def update_stop_loss(symbol, new_sl, side):
    """更新止損：取消舊的 algo STOP_MARKET，掛新的"""
    global client
    tick_size, _, _ = get_symbol_info(symbol)

    try:
        # 查詢現有的 algo stop orders
        algo_orders = client.sign_request("GET", "/fapi/v1/openAlgoOrders", {
            "symbol": symbol,
            "algoType": "CONDITIONAL",
        })
        for order in algo_orders:
            if order.get("orderType") == "STOP_MARKET" and order.get("algoStatus") == "NEW":
                algo_id = order["algoId"]
                client.sign_request("DELETE", "/fapi/v1/algoOrder", {"algoId": algo_id})
                print(f"  Cancelled old algo SL {algo_id}")

        # 掛新的
        entry_side = "BUY" if side == "long" else "SELL"
        _place_stop_order(symbol, entry_side, 0, new_sl, "STOP_MARKET", tick_size)
    except Exception as e:
        print(f"update_stop_loss error: {e}")
        client = new_session()


# ── Excel 紀錄 ───────────────────────────────────────────────
def log_pnl_to_xlsx_trade_record(records: list):
    filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_pnl_log.xlsx")
    headers = ["交易對", "工具", "平倉價格", "訂單數量", "交易類型", "已結盈虧", "成交時間"]

    try:
        file_exists = os.path.exists(filename)
        if not file_exists:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            existing_rows = set()
        else:
            wb = load_workbook(filename)
            ws = wb.active
            existing_rows = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                unique_key = (row[0], row[2], row[3], row[4], row[6])
                existing_rows.add(unique_key)

        insert_count = 0
        for record in records:
            unique_key = (
                record["symbol"], record["exit_price"],
                record["qty"], record["side"], record["close_time"]
            )
            if unique_key in existing_rows:
                continue
            ws.append([
                record["symbol"], "USDT Perpetual",
                record["exit_price"], record["qty"],
                record["side"], record["pnl"], record["close_time"]
            ])
            insert_count += 1

        wb.save(filename)
        wb.close()

        if insert_count > 0:
            wb = load_workbook(filename)
            ws = wb.active
            pnl_summary = {}
            win_count = lose_count = total_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                sym = row[0]
                pnl = float(row[5]) if row[5] is not None else 0.0
                pnl_summary[sym] = pnl_summary.get(sym, 0.0) + pnl
                if pnl > 0:
                    win_count += 1
                else:
                    lose_count += 1
                total_count += 1
            wb.close()

            win_rate = (win_count / total_count * 100) if total_count > 0 else 0.0
            summary_lines = [f"累計已結盈虧:"]
            for sym, total_pnl in pnl_summary.items():
                summary_lines.append(f"  {sym}: {total_pnl:.2f} USDT")
            summary_lines.append(f"勝率: {win_count}W {lose_count}L = {win_rate:.1f}%")

            msg = f"新交易紀錄 {insert_count} 筆\n" + "\n".join(summary_lines)
            print(msg)
            send_telegram_message(msg)

    except Exception as e:
        print(f"log_pnl_to_xlsx error: {e}")


def record_trade(symbol=None):
    """從 Binance 撈最近平倉紀錄"""
    symbol = symbol or SYMBOL
    global client
    try:
        trades = client.get_account_trades(symbol=symbol, limit=50)
        now_ms = int(time.time() * 1000)
        one_hour_ago = now_ms - 3600 * 1000

        records = []
        for t in trades:
            if int(t["time"]) < one_hour_ago:
                continue
            records.append({
                "symbol": symbol,
                "exit_price": float(t["price"]),
                "qty": float(t["qty"]),
                "side": t["side"],
                "pnl": float(t.get("realizedPnl", 0)),
                "close_time": datetime.fromtimestamp(int(t["time"]) / 1000).strftime("%Y-%m-%d %H:%M:%S"),
            })

        if records:
            log_pnl_to_xlsx_trade_record(records)

    except Exception as e:
        print(f"record_trade error: {e}")
        client = new_session()


# ── 測試入口 ─────────────────────────────────────────────────
if __name__ == "__main__":
    env = "TESTNET" if is_testnet else "PRODUCTION"
    print(f"=== Binance Futures ({env}) ===")
    print(f"Symbol: {SYMBOL}")

    # 測試連線
    balance = get_available_balance()
    print(f"Balance: {balance:.2f} USDT")

    # 測試商品規格
    tick, step, minq = get_symbol_info()
    print(f"Tick: {tick}, Step: {step}, MinQty: {minq}")

    # 測試持倉查詢
    positions = get_positions()
    if positions:
        for p in positions:
            print(f"Position: {p['side']} {p['size']} @ {p['entry_price']:.2f}")
    else:
        print("No open positions")

    # 設定槓桿
    set_leverage()

    print("\nReady. To test order: place_order(SYMBOL, 'BUY')")
