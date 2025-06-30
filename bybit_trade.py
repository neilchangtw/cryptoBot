import os
import time
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from pybit.unified_trading import HTTP
from telegram_notify import send_telegram_message

# 讀取 .env 環境變數
load_dotenv()

# 讀取 API 金鑰設定
api_key = os.getenv("BYBIT_API_KEY")
api_secret = os.getenv("BYBIT_API_SECRET")
BYBIT_BASE_URL = os.getenv("BYBIT_BASE_URL", "https://api.bybit.com")

# 倉位與風控參數
fixed_amount = float(os.getenv("FIXED_AMOUNT", 100))
leverage = float(os.getenv("LEVERAGE", 20))
max_order_amount = float(os.getenv("MAX_ORDER_AMOUNT", 0))  # 0 代表不限制
cooldown_seconds = int(os.getenv("COOLDOWN_SECONDS", 600))

# 建立 Bybit Session
def new_session():
    return HTTP(
        api_key=api_key,
        api_secret=api_secret,
        testnet=False,
        demo=True,
        recv_window=10000
    )

session = new_session()
last_trade_time = {}

# 查詢帳戶可用資金
def get_available_balance():
    global session
    try:
        result = session.get_wallet_balance(accountType="UNIFIED")
        usdt_balance = float(result["result"]["list"][0]["totalEquity"])
        return usdt_balance
    except Exception as e:
        print("❌ 查詢帳戶餘額失敗:", e)
        send_telegram_message(f"❗查詢帳戶餘額失敗: {e}")
        session = new_session()
        return 0.0

# 查詢商品規格
def get_symbol_info(symbol):
    global session
    try:
        res = session.get_instruments_info(category="linear", symbol=symbol)
        info = res["result"]["list"][0]
        tick_size = float(info["priceFilter"]["tickSize"])
        qty_step = float(info["lotSizeFilter"]["qtyStep"])
        min_qty = float(info["lotSizeFilter"]["minOrderQty"])
        return tick_size, qty_step, min_qty
    except Exception as e:
        print("❌ 查詢商品資訊失敗:", e)
        send_telegram_message(f"❗查詢商品資訊失敗: {e}")
        session = new_session()
        return 0.01, 0.001, 0.001

def round_to_tick(price, tick_size):
    return round(round(price / tick_size) * tick_size, 8)

def round_to_lot(qty, qty_step, min_qty):
    qty = round(round(qty / qty_step) * qty_step, 8)
    return max(qty, min_qty)

def place_order(symbol, side, price, stop_loss=None, take_profit=None, strategy_id="default"):
    global session, last_trade_time

    now = time.time()
    tick_size, qty_step, min_qty = get_symbol_info(symbol)

    # === 模擬平倉邏輯（直接反向市價單平倉） ===
    if side.upper() in ["BUY", "SELL"] and stop_loss is None and take_profit is None:
        try:
            pos_info = session.get_positions(category="linear", symbol=symbol)["result"]["list"][0]
            current_side = pos_info["side"]
            qty = float(pos_info["size"])

            if qty <= 0 or (
                    (side.upper() == "BUY" and current_side != "Sell") or
                    (side.upper() == "SELL" and current_side != "Buy")
            ):
                send_telegram_message(f"⚠️ 無需Exit平倉：{symbol} 當前倉位不符或已無部位")
                return

            res = session.place_order(
                category="linear",
                symbol=symbol,
                side=side.capitalize(),
                orderType="Market",
                qty=str(qty),
                timeInForce="IOC"
            )
            # 嘗試抓回傳成交價（以 bybit v5 unified_trading 套件為例）
            try:
                exec_price = res["result"]["avgPrice"] if "avgPrice" in res["result"] else res["result"].get("orderPrice")
            except Exception:
                exec_price = None

            send_telegram_message(
                f"📤 Exit平倉成功: {symbol}，原持倉: {current_side}，平倉方向: {side.upper()}，數量: {qty}\n平倉價格: {exec_price}"
            )
            print(f"✅ Exit平倉成功: {res}")
            last_trade_time[(strategy_id, symbol)] = now
            record_trade(symbol)

        except Exception as e:
            print("❌ Exit平倉失敗:", e)
            send_telegram_message(f"❌ Exit平倉失敗: {e}")
            session = new_session()
        return   # <-- 這裡有 return，是正確的

    # 真實平倉 CLOSE 支援
    if side.upper() == "CLOSE":
        try:
            pos_info = session.get_positions(category="linear", symbol=symbol)["result"]["list"][0]
            side_pos = pos_info["side"]
            qty = float(pos_info["size"])

            if qty <= 0:
                send_telegram_message(f"⚠️ 無需平倉：{symbol} 無未平倉部位")
                return

            opposite_side = "Sell" if side_pos == "Buy" else "Buy"

            res = session.place_order(
                category="linear",
                symbol=symbol,
                side=opposite_side,
                orderType="Market",
                qty=str(qty),
                timeInForce="IOC"
            )

            send_telegram_message(f"📤 已平倉 {symbol}，方向: {opposite_side}，數量: {qty}")
            print(f"✅ 平倉成功: {res}")
            last_trade_time[(strategy_id, symbol)] = now
            record_trade(symbol)

        except Exception as e:
            print("❌ 平倉失敗:", e)
            send_telegram_message(f"❌ 平倉失敗: {e}")
            session = new_session()
        return

    cooldown_key = (strategy_id, symbol)
    if cooldown_key in last_trade_time and now - last_trade_time[cooldown_key] < cooldown_seconds:
        print(f"⏳ 冷卻中（策略: {strategy_id}, 幣種: {symbol}），避免頻繁下單")
        return

    price = round_to_tick(price, tick_size)
    balance = get_available_balance()

    total_usd = min(fixed_amount * leverage, max_order_amount) if max_order_amount > 0 else fixed_amount * leverage

    qty = total_usd / price
    qty = round_to_lot(qty, qty_step, min_qty)

    if qty < min_qty:
        send_telegram_message(f"❌ 下單失敗：數量 {qty} 低於最小下單量 {min_qty}")
        return

    sl_price = round_to_tick(float(stop_loss), tick_size) if stop_loss else None
    tp_price = round_to_tick(float(take_profit), tick_size) if take_profit else None

    try:
        params = {
            "category": "linear",
            "symbol": symbol,
            "side": side.capitalize(),
            "orderType": "Market",
            "qty": str(qty),
            "timeInForce": "IOC",
        }
        if sl_price:
            params["stopLoss"] = str(sl_price)
        if tp_price:
            params["takeProfit"] = str(tp_price)

        res = session.place_order(**params)

        print(f"✅ {side} 成功下單: {res}")
        send_telegram_message(
            f"✅ 已市價 {side} {symbol}\n數量: {qty}\n價格: {price}\n止損: {sl_price}\n止盈: {tp_price}\n總倉位: {total_usd} USDT"
        )
        last_trade_time[cooldown_key] = now
        record_trade(symbol)

    except Exception as e:
        print("❌ 下單失敗:", e)
        send_telegram_message(f"❌ 下單失敗: {e}")
        session = new_session()

# Excel 紀錄部分
def log_pnl_to_xlsx_trade_record(records: list):
    filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_pnl_log.xlsx")
    headers = ["交易對", "工具", "平倉價格", "訂單數量", "交易類型", "已結盈虧", "成交時間"]

    try:
        file_exists = os.path.exists(filename)
        if not file_exists:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            existing_rows = set()
        else:
            wb = load_workbook(filename)
            ws = wb.active
            existing_rows = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                unique_key = (row[0], row[2], row[3], row[4], row[6])
                existing_rows.add(unique_key)

        insert_count = 0

        for record in records:
            unique_key = (
                record["symbol"],
                record["exit_price"],
                record["qty"],
                record["side"],
                record["close_time"]
            )
            if unique_key in existing_rows:
                continue
            ws.append([
                record["symbol"],
                "USDT 永續",
                record["exit_price"],
                record["qty"],
                record["side"],
                record["pnl"],
                record["close_time"]
            ])
            insert_count += 1

        wb.save(filename)
        wb.close()

        if insert_count > 0:
            # 重新讀取並統計每個幣種的總結盈虧
            wb = load_workbook(filename)
            ws = wb.active
            pnl_summary = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                symbol = row[0]
                pnl = float(row[5]) if row[5] is not None else 0.0
                pnl_summary[symbol] = pnl_summary.get(symbol, 0.0) + pnl
            wb.close()

            summary_lines = [f"📊 累計已結盈虧："]
            for sym, total_pnl in pnl_summary.items():
                emoji = "💰" if total_pnl >= 0 else "🔻"
                summary_lines.append(f"{emoji} {sym}: {total_pnl:.2f} USDT")

            summary_msg = (
                    f"📗 新交易紀錄寫入 {insert_count} 筆\n"
                    + "\n".join(summary_lines)
            )
            print(summary_msg)
            send_telegram_message(summary_msg)
        else:
            print("📗 無新交易紀錄，跳過通知")

    except Exception as e:
        print("❌ 寫入 XLSX 失敗：", e)
        send_telegram_message(f"❗寫入交易紀錄失敗：{e}")


def record_trade(symbol):
    global session
    try:
        now_ts = int(datetime.utcnow().timestamp() * 1000)
        one_hour_ago_ts = now_ts - 1 * 60 * 60 * 1000

        result = session.get_closed_pnl(category="linear", symbol=symbol, limit=100)
        closed_records = result["result"]["list"]
        trade_records = []

        for record in closed_records:
            updated_time = int(record["updatedTime"])
            if updated_time < one_hour_ago_ts:
                continue

            pnl = float(record["closedPnl"])
            qty = float(record["qty"])
            exit_price = float(record["avgExitPrice"])
            side = record["side"]
            close_time = datetime.fromtimestamp(updated_time / 1000).strftime("%Y-%m-%d %H:%M:%S")

            trade_records.append({
                "symbol": symbol,
                "exit_price": exit_price,
                "qty": qty,
                "side": side,
                "pnl": pnl,
                "close_time": close_time
            })

        if trade_records:
            log_pnl_to_xlsx_trade_record(trade_records)
        else:
            msg = f"⚠️ 無平倉紀錄：{symbol} 最近 1 小時內無平倉資料"
            print(msg)
            # send_telegram_message(msg)

    except Exception as e:
        print("❌ 撈取平倉紀錄失敗：", e)
        send_telegram_message(f"❗平倉紀錄失敗: {e}")
        session = new_session()